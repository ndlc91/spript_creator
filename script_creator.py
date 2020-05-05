# The purpose of this app is to generate scripts using various info in an xcel spreadsheet
# The app needs to pull info from cells in the array, and then generate a few scripts by looping through a few lists to pick the right script intro and pitch


import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.load_workbook('days2.xlsx')
sheet = wb['straight-thru-references']


# Script variables
intro_counter = 0
pitches_counter = 0

for i in range(2, 93):  # iterate through every row
    day = sheet.cell(row=i, column=1).value
    reference = sheet.cell(row=i, column=3).value
    # change the intro by cycling through the intro_counter
    intros = [f'Hey there!  And welcome to Day {day} of the Bible in a year podcast!',
          f'Welcome to Day {day} of the Bible in a year podcast!', f'God bless!  And welcome to Day {day} of The Bible in a year podcast!']

    pitches = ['If you enjoy today’s Podcast, consider checking out our App. There you can listen to the entire Bible, explore new listening plans, and use Dwell Mode to memorize or meditate on your favorite verses. We hope you enjoy today’s reading!', 'We hope you enjoy todays reading!', 'We hope you enjoy todays reading!', 'We hope you enjoy todays reading!', 'If you have been enjoying this Podcast, consider checking out our App.  In the app, you can create your ideal listening experience by picking the voice, music, translation, and listening speed as you listen to the Bible.  We hope you enjoy today’s reading!”',
           'We hope you enjoy todays reading!', 'We hope you enjoy todays reading!', 'We hope you enjoy todays reading!', 'If you have been enjoying this Podcast, consider checking out our App.  There you can pick from hundreds of listening plans, playlists, and passages that will help you explore the Bible in a new way.  We hope you enjoy today’s reading!”', 'We hope you enjoy todays reading!', 'We hope you enjoy todays reading!', 'We hope you enjoy todays reading!']

    reference = f' Our reading today is {reference}. '

    if intro_counter == 0:
        intro = intros[intro_counter]
        intro_counter = intro_counter + 1
        print()
    elif intro_counter == 1:
        intro = intros[intro_counter]
        intro_counter = intro_counter + 1
    elif intro_counter == 2:
        intro = intros[intro_counter]
        intro_counter = 0

    # change the pitch by cycling through pitches_counter
    
    if pitches_counter == 0:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 1:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 2:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 3:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 4:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 5:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 6:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 7:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 8:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 9:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 10:
        pitch = pitches[pitches_counter]
        pitches_counter = pitches_counter + 1
    elif pitches_counter == 11:
        pitch = pitches[pitches_counter]
        pitches_counter = 0
    

    sheet.cell(row=i, column=4).value = (intro + reference + pitch)

wb.save("/Users/nathandelacruz/School/PythonProjects/Automate/WorkingWithSpreadsheets/days2_filled.xlsx")